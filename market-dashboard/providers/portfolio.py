from __future__ import annotations

import hashlib
import math
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

try:
    import akshare as ak
except Exception:  # noqa: BLE001
    ak = None


LISTED_ETF_CODES = {"560090", "513070", "513260", "513050"}
BOND_KEYWORDS = ("债", "短债", "可转债", "收益债", "中短债", "高收益债")
CASH_KEYWORDS = ("货币", "现金", "逆回购")
LINKED_KEYWORDS = ("ETF联接", "联接发起式")
EQUITY_KEYWORDS = ("股票", "混合", "成长精选", "优质企业")
QDII_KEYWORDS = ("QDII", "全球", "海外", "纳指", "越南", "标普")
FUNDISH_KEYWORDS = ("基金", "指数", "LOF", "ETF")


@dataclass
class PortfolioWorkbook:
    positions: list[dict[str, Any]]
    closed_positions: list[dict[str, Any]]
    transactions: list[dict[str, Any]]
    summary: dict[str, Any]


def source_file_hash(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def make_batch_id(path: str | Path, as_of_date: date) -> str:
    return f"portfolio-{as_of_date.isoformat()}-{source_file_hash(path)[:12]}"


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text in {"", "-", "--", "None", "nan"}:
            return None
        return text
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def to_date(value: Any) -> date | None:
    value = clean_value(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


DATE_LABEL_KEYWORDS = ("持仓日期", "持仓日", "导出日期", "导出时间", "统计日期", "截止日期", "数据日期", "资产日期", "净值日期")
DATE_TEXT_RE = re.compile(r"(?<!\d)(20\d{2})[-./年]?([01]?\d)[-./月]?([0-3]?\d)(?:日)?(?!\d)")


def _plausible_as_of_date(value: date | None) -> date | None:
    if value is None or value.year < 2020 or value.year > 2100:
        return None
    return value


def _date_from_text(text: str) -> date | None:
    for year, month, day in DATE_TEXT_RE.findall(text):
        try:
            return _plausible_as_of_date(date(int(year), int(month), int(day)))
        except ValueError:
            continue
    return None


def infer_as_of_date(path: str | Path, filename: str | None = None) -> date | None:
    """Infer the portfolio snapshot date without trusting transaction history dates."""
    filename_date = _date_from_text(filename or Path(path).name)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True, max_row=40)
            for row in rows:
                values = list(row[:20])
                for index, value in enumerate(values):
                    text = str(clean_value(value) or "")
                    if not text or not any(keyword in text for keyword in DATE_LABEL_KEYWORDS):
                        continue
                    direct = _date_from_text(text) or _plausible_as_of_date(to_date(value))
                    if direct:
                        return direct
                    nearby_values = values[index + 1:index + 4] + values[max(0, index - 2):index]
                    for nearby in nearby_values:
                        candidate = _date_from_text(str(clean_value(nearby) or "")) or _plausible_as_of_date(to_date(nearby))
                        if candidate:
                            return candidate
    finally:
        workbook.close()
    return filename_date


def to_time(value: Any) -> time | None:
    value = clean_value(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return None


def to_decimal(value: Any) -> Decimal | None:
    value = clean_value(value)
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except Exception:  # noqa: BLE001
        return None


def to_float(value: Any) -> float | None:
    value = clean_value(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_int(value: Any) -> int | None:
    value = clean_value(value)
    if value is None:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def json_safe(value: Any) -> Any:
    value = clean_value(value)
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [json_safe(item) for item in value]
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "item"):
        try:
            return json_safe(value.item())
        except Exception:  # noqa: BLE001
            pass
    return value


def raw_json_safe(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in row.items():
        value = clean_value(value)
        out[key] = json_safe(value)
    return out


def read_sheet(path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        records: list[dict[str, Any]] = []
        for values in rows:
            row = {headers[index]: clean_value(value) for index, value in enumerate(values) if headers[index]}
            if any(value is not None for value in row.values()):
                records.append(row)
        return records
    finally:
        workbook.close()


def read_portfolio_workbook(path: str | Path, as_of_date: date) -> PortfolioWorkbook:
    position_rows = read_sheet(path, "持仓数据")
    closed_rows = read_sheet(path, "已清仓")
    transaction_rows = read_sheet(path, "交易记录")

    summary: dict[str, Any] = {"as_of_date": as_of_date.isoformat()}
    positions: list[dict[str, Any]] = []
    for row in position_rows:
        code = str(row.get("代码") or "").strip()
        if code == "汇总":
            summary = raw_json_safe(row) | summary
            continue
        if not code:
            continue
        positions.append(raw_json_safe(row))

    return PortfolioWorkbook(
        positions=positions,
        closed_positions=[raw_json_safe(row) for row in closed_rows],
        transactions=[raw_json_safe(row) for row in transaction_rows],
        summary=summary,
    )


def classify_security(code: str, name: str, fund_catalog: set[str] | None = None) -> tuple[str, str | None]:
    fund_catalog = fund_catalog or set()
    code = str(code or "").zfill(6)
    name = name or ""
    upper_name = name.upper()

    if code in LISTED_ETF_CODES:
        return "etf_listed", "CN"
    if any(keyword in name for keyword in CASH_KEYWORDS) or code in {"511990"}:
        return "money_or_cash", "CN"
    if any(keyword in name for keyword in BOND_KEYWORDS):
        return "bond_fund", "CN"
    if any(keyword in name for keyword in LINKED_KEYWORDS):
        return "fund_linked", "CN"
    if any(keyword in upper_name for keyword in QDII_KEYWORDS):
        return "qdii_fund", "GLOBAL"
    if any(keyword in name for keyword in EQUITY_KEYWORDS):
        return "equity_fund", "CN"
    if any(keyword in upper_name for keyword in FUNDISH_KEYWORDS):
        return "other_fund", "CN"
    if re.fullmatch(r"\d{6}", code) and code not in fund_catalog:
        return "stock_a", "CN"
    if re.fullmatch(r"\d{6}", code):
        return "other_fund", "CN"
    return "other_fund", None


def fund_catalog() -> set[str]:
    """Return an empty set — classification is done via heuristic rules + per-code checks."""
    return set()


def _first_existing(record: dict[str, Any], names: tuple[str, ...]) -> Any:
    for name in names:
        if name in record and clean_value(record[name]) is not None:
            return clean_value(record[name])
    return None


def _df_records(df: pd.DataFrame | None) -> list[dict[str, Any]]:
    if df is None or df.empty:
        return []
    return [raw_json_safe(row) for row in df.to_dict(orient="records")]


def _latest_period(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not records:
        return []
    period_keys = ("报告期", "季度", "截止日期", "净值日期", "日期")
    periods = [_first_existing(row, period_keys) for row in records]
    latest = max((str(period) for period in periods if period is not None), default="")
    if not latest:
        return records[:10]
    return [row for row in records if str(_first_existing(row, period_keys) or "") == latest][:10]


def _parse_weight(value: Any) -> float | None:
    value = clean_value(value)
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace("%", "").strip()
    number = to_float(value)
    if number is None:
        return None
    return number / 100 if abs(number) > 1 else number


def fetch_fund_metadata(fund_code: str) -> dict[str, Any]:
    if ak is None:
        return {"raw_json": {"error": "akshare is not available"}}
    try:
        df = ak.fund_individual_basic_info_xq(symbol=fund_code)
        records = _df_records(df)
        flat: dict[str, Any] = {}
        for row in records:
            key = _first_existing(row, ("item", "项目", "指标", "name"))
            value = _first_existing(row, ("value", "内容", "值", "数据"))
            if key:
                flat[str(key)] = value
        merged = flat or (records[0] if records else {})
        size_value = _first_existing(merged, ("基金规模", "资产规模", "最新规模"))
        size = None
        size_unit = None
        if size_value is not None:
            match = re.search(r"([0-9.]+)\s*([^0-9.]*)", str(size_value))
            if match:
                size = to_decimal(match.group(1))
                size_unit = match.group(2).strip() or None
        return {
            "fund_type": _first_existing(merged, ("基金类型", "类型")),
            "manager_name": _first_existing(merged, ("基金经理", "基金经理人", "经理")),
            "fund_size": size,
            "size_unit": size_unit,
            "management_fee": _parse_weight(_first_existing(merged, ("管理费率", "管理费"))),
            "custody_fee": _parse_weight(_first_existing(merged, ("托管费率", "托管费"))),
            "sales_service_fee": _parse_weight(_first_existing(merged, ("销售服务费率", "销售服务费"))),
            "purchase_fee": _parse_weight(_first_existing(merged, ("最高申购费率", "申购费率", "买入费率"))),
            "redemption_fee": _first_existing(merged, ("赎回费率", "卖出费率")),
            "inception_date": to_date(_first_existing(merged, ("成立日期", "基金成立日"))),
            "benchmark": _first_existing(merged, ("业绩比较基准", "跟踪标的")),
            "raw_json": {"records": records, "flat": raw_json_safe(merged)},
        }
    except Exception as exc:  # noqa: BLE001
        return {"raw_json": {"error": str(exc)}}


def _fetch_yearly_records(callable_name: str, fund_code: str, years: list[str]) -> list[dict[str, Any]]:
    if ak is None:
        return [{"error": "akshare is not available"}]
    records: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    func = getattr(ak, callable_name)
    for year in years:
        try:
            df = func(symbol=fund_code, date=year)
            records.extend(_df_records(df))
        except Exception as exc:  # noqa: BLE001
            errors.append({"year": year, "error": str(exc)})
    if errors and not records:
        return [{"error": errors}]
    return records


def normalize_fund_holding(row: dict[str, Any], underlying_type: str) -> dict[str, Any]:
    return {
        "underlying_code": str(_first_existing(row, ("股票代码", "债券代码", "代码", "证券代码")) or "").zfill(6),
        "underlying_name": _first_existing(row, ("股票名称", "债券名称", "名称", "证券简称")) or "",
        "underlying_type": underlying_type,
        "report_period": str(_first_existing(row, ("报告期", "季度", "截止日期")) or ""),
        "holding_rank": to_int(_first_existing(row, ("序号", "排名", "持仓排名"))),
        "holding_weight_in_parent": _parse_weight(_first_existing(row, ("占净值比例", "持仓占比", "占基金净值比例", "比例"))),
        "shares": to_decimal(_first_existing(row, ("持股数", "持债数量", "数量"))),
        "market_value": to_decimal(_first_existing(row, ("持仓市值", "市值"))),
        "raw_json": raw_json_safe(row),
    }


def fetch_fund_stock_holdings(fund_code: str, years: list[str]) -> list[dict[str, Any]]:
    records = _fetch_yearly_records("fund_portfolio_hold_em", fund_code, years)
    if records and "error" in records[0]:
        return [{"raw_json": records[0]}]
    return [normalize_fund_holding(row, "stock") for row in _latest_period(records)]


def fetch_fund_bond_holdings(fund_code: str, years: list[str]) -> list[dict[str, Any]]:
    records = _fetch_yearly_records("fund_portfolio_bond_hold_em", fund_code, years)
    if records and "error" in records[0]:
        return [{"raw_json": records[0]}]
    return [normalize_fund_holding(row, "bond") for row in _latest_period(records)]


def fetch_fund_industry_allocation(fund_code: str, years: list[str]) -> list[dict[str, Any]]:
    records = _fetch_yearly_records("fund_portfolio_industry_allocation_em", fund_code, years)
    if records and "error" in records[0]:
        return [{"raw_json": records[0]}]
    rows: list[dict[str, Any]] = []
    for row in _latest_period(records):
        rows.append(
            {
                "industry_name": _first_existing(row, ("行业类别", "行业名称", "名称")) or "未知",
                "report_period": str(_first_existing(row, ("报告期", "季度", "截止日期")) or ""),
                "weight_in_parent": _parse_weight(_first_existing(row, ("占净值比例", "比例", "持仓占比"))),
                "raw_json": raw_json_safe(row),
            }
        )
    return rows


def fetch_open_fund_nav(fund_code: str) -> pd.DataFrame:
    if ak is None:
        return pd.DataFrame()
    try:
        return ak.fund_open_fund_info_em(symbol=fund_code, indicator="单位净值走势", period="1年")
    except Exception:  # noqa: BLE001
        return pd.DataFrame()


def fetch_price_history(code: str, security_type: str, start_date: str, end_date: str) -> pd.DataFrame:
    if ak is None:
        return pd.DataFrame()
    try:
        if security_type == "etf_listed":
            return ak.fund_etf_hist_em(symbol=code, period="daily", start_date=start_date, end_date=end_date, adjust="qfq")
        if security_type == "stock_a":
            return ak.stock_zh_a_hist(symbol=code, period="daily", start_date=start_date, end_date=end_date, adjust="qfq")
    except Exception:  # noqa: BLE001
        return pd.DataFrame()
    return pd.DataFrame()


def compute_returns_and_drawdown(nav_df: pd.DataFrame) -> dict[str, Any]:
    if nav_df is None or nav_df.empty:
        return {}
    date_col = next((col for col in ("净值日期", "日期", "date") if col in nav_df.columns), nav_df.columns[0])
    value_col = next((col for col in ("单位净值", "收盘", "close", "累计净值") if col in nav_df.columns), None)
    if value_col is None:
        numeric_cols = [col for col in nav_df.columns if pd.api.types.is_numeric_dtype(nav_df[col])]
        value_col = numeric_cols[0] if numeric_cols else None
    if value_col is None:
        return {}
    df = nav_df[[date_col, value_col]].copy()
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df[value_col] = pd.to_numeric(df[value_col], errors="coerce")
    df = df.dropna().sort_values(date_col)
    if df.empty:
        return {}
    series = df[value_col]

    def period_return(days: int) -> float | None:
        if len(series) < 2:
            return None
        cutoff = df[date_col].iloc[-1] - pd.Timedelta(days=days)
        base = df[df[date_col] <= cutoff]
        base_value = base[value_col].iloc[-1] if not base.empty else series.iloc[0]
        return float(series.iloc[-1] / base_value - 1) if base_value else None

    drawdown = (series / series.cummax() - 1).min()
    return {
        "return_1m": period_return(30),
        "return_3m": period_return(90),
        "return_6m": period_return(180),
        "return_1y": period_return(365),
        "max_drawdown_1y": float(drawdown),
        "nav_latest": to_decimal(series.iloc[-1]),
        "nav_date": df[date_col].iloc[-1].date(),
    }


def build_direct_or_fallback_underlying(position: dict[str, Any], underlying_type: str, status: str) -> dict[str, Any]:
    amount = to_decimal(position.get("持有金额")) or Decimal("0")
    weight = to_float(position.get("仓位占比")) or 0.0
    return {
        "parent_code": str(position.get("代码")).zfill(6),
        "parent_name": position.get("名称") or "",
        "parent_type": position.get("security_type") or underlying_type,
        "underlying_code": str(position.get("代码")).zfill(6),
        "underlying_name": position.get("名称") or "",
        "underlying_type": underlying_type,
        "report_period": "",
        "holding_rank": 1,
        "holding_weight_in_parent": 1.0,
        "parent_portfolio_weight": weight,
        "lookthrough_portfolio_weight": weight,
        "lookthrough_amount": amount,
        "shares": to_decimal(position.get("持有数量")),
        "market_value": amount,
        "source": "excel",
        "raw_json": {"lookthrough_status": status},
    }


def enrich_fund_underlying(position: dict[str, Any], holdings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    parent_weight = to_float(position.get("仓位占比")) or 0.0
    parent_amount = to_decimal(position.get("持有金额")) or Decimal("0")
    out: list[dict[str, Any]] = []
    for row in holdings:
        weight_in_parent = row.get("holding_weight_in_parent")
        if weight_in_parent is None:
            continue
        out.append(
            {
                "parent_code": str(position.get("代码")).zfill(6),
                "parent_name": position.get("名称") or "",
                "parent_type": position.get("security_type") or "other_fund",
                "underlying_code": row.get("underlying_code") or "",
                "underlying_name": row.get("underlying_name") or "",
                "underlying_type": row.get("underlying_type") or "unknown",
                "report_period": row.get("report_period") or "",
                "holding_rank": row.get("holding_rank"),
                "holding_weight_in_parent": weight_in_parent,
                "parent_portfolio_weight": parent_weight,
                "lookthrough_portfolio_weight": parent_weight * weight_in_parent,
                "lookthrough_amount": parent_amount * Decimal(str(weight_in_parent)),
                "shares": row.get("shares"),
                "market_value": row.get("market_value"),
                "source": "akshare",
                "raw_json": row.get("raw_json") or {},
            }
        )
    return out


def compute_portfolio_allocations(positions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets = {
        "股票/股票基金/ETF": {"types": {"stock_a", "etf_listed", "fund_linked", "equity_fund"}, "amount": Decimal("0"), "weight": 0.0},
        "债券基金": {"types": {"bond_fund"}, "amount": Decimal("0"), "weight": 0.0},
        "货币/现金/逆回购": {"types": {"money_or_cash"}, "amount": Decimal("0"), "weight": 0.0},
        "QDII": {"types": {"qdii_fund"}, "amount": Decimal("0"), "weight": 0.0},
        "其他": {"types": {"other_fund"}, "amount": Decimal("0"), "weight": 0.0},
    }
    total_amount = Decimal("0")
    total_weight = 0.0
    for position in positions:
        amount = to_decimal(position.get("持有金额")) or Decimal("0")
        weight = to_float(position.get("仓位占比")) or 0.0
        total_amount += amount
        total_weight += weight
        security_type = position.get("security_type")
        bucket_name = next((name for name, bucket in buckets.items() if security_type in bucket["types"]), "其他")
        buckets[bucket_name]["amount"] += amount
        buckets[bucket_name]["weight"] += weight

    if 0 < total_weight < 0.999:
        estimated_total = total_amount / Decimal(str(total_weight))
        cash_amount = estimated_total - total_amount
        buckets["货币/现金/逆回购"]["amount"] += cash_amount
        buckets["货币/现金/逆回购"]["weight"] += max(0.0, 1.0 - total_weight)

    return [
        {
            "allocation_bucket": name,
            "amount": bucket["amount"],
            "weight": bucket["weight"],
            "source": "computed",
            "raw_json": {"estimated_cash": name == "货币/现金/逆回购" and total_weight < 0.999},
        }
        for name, bucket in buckets.items()
        if bucket["amount"] or bucket["weight"]
    ]


def compute_risk_metrics(
    positions: list[dict[str, Any]],
    underlying_holdings: list[dict[str, Any]],
    fund_metadata: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    weights = sorted((to_float(row.get("仓位占比")) or 0.0 for row in positions), reverse=True)
    total_amount = sum((to_decimal(row.get("持有金额")) or Decimal("0") for row in positions), Decimal("0"))
    total_weight = sum(weights)
    equity_types = {"stock_a", "etf_listed", "fund_linked", "equity_fund"}
    bond_types = {"bond_fund"}
    qdii_types = {"qdii_fund"}

    def type_weight(types: set[str]) -> float:
        return sum(to_float(row.get("仓位占比")) or 0.0 for row in positions if row.get("security_type") in types)

    underlying_weights = sorted((row.get("lookthrough_portfolio_weight") or 0.0 for row in underlying_holdings), reverse=True)
    max_drawdowns = [
        item.get("max_drawdown_1y")
        for item in fund_metadata
        if item.get("max_drawdown_1y") is not None
    ]
    metrics = {
        "total_assets": float(total_amount) if total_amount is not None else None,
        "cash_weight_estimated": max(0.0, 1.0 - total_weight) if total_weight < 1 else 0.0,
        "position_count": float(len(positions)),
        "top3_position_weight": sum(weights[:3]),
        "top5_position_weight": sum(weights[:5]),
        "top10_underlying_weight": sum(underlying_weights[:10]),
        "equity_like_weight": type_weight(equity_types),
        "bond_like_weight": type_weight(bond_types),
        "qdii_weight": type_weight(qdii_types),
        "portfolio_profit_loss": float(sum((to_decimal(row.get("持有盈亏")) or Decimal("0") for row in positions), Decimal("0"))),
        "portfolio_cumulative_profit_loss": float(sum((to_decimal(row.get("累计盈亏")) or Decimal("0") for row in positions), Decimal("0"))),
        "max_drawdown_1y": min(max_drawdowns) if max_drawdowns else None,
    }
    return [
        {
            "metric_scope": "portfolio",
            "subject_code": "PORTFOLIO",
            "subject_name": "组合",
            "metric_name": name,
            "metric_value": value,
            "metric_unit": "ratio" if name.endswith("weight") or name.endswith("drawdown_1y") else None,
            "source": "computed",
            "raw_json": {},
        }
        for name, value in metrics.items()
    ]
